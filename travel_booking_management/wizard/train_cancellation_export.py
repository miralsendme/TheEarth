# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, _

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None


class TrainCancellationExport(models.TransientModel):
    _name = 'travel.train.cancellation.export'
    _description = 'Export Train Cancellations to Excel'

    date_from = fields.Date(string='From Date')
    date_to = fields.Date(string='To Date')
    state_filter = fields.Selection([
        ('all', 'All'), ('draft', 'Draft'), ('confirmed', 'Confirmed'),
        ('refunded', 'Refunded'), ('rejected', 'Rejected'),
    ], string='Status', default='all')
    excel_file = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='File Name', readonly=True)

    def action_export(self):
        self.ensure_one()
        domain = []
        if self.date_from:
            domain.append(('cancellation_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('cancellation_date', '<=', self.date_to))
        if self.state_filter and self.state_filter != 'all':
            domain.append(('state', '=', self.state_filter))
        if self.env.context.get('active_ids') and self.env.context.get('active_model') == 'travel.train.cancellation':
            domain = [('id', 'in', self.env.context['active_ids'])]
            if self.state_filter and self.state_filter != 'all':
                domain.append(('state', '=', self.state_filter))

        records = self.env['travel.train.cancellation'].search(domain, order='cancellation_date asc')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Train Cancellations'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = [
            'Cancellation Ref', 'Billing Company', 'Cancellation Date', 'Travel Date',
            'Booking Executive', 'Employee Code', 'Doc No/Req By',
            'Passenger(s)', 'No. of Passengers', 'Origin', 'Destination',
            'Quota', 'PNR Number', 'Train Number',
            'Total Fare', 'Service Charge', 'Refund Amount',
            'Mode of Refund', 'Confirmed By', 'Status', 'Remarks',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, rc in enumerate(records, 2):
            values = [
                rc.name or '',
                rc.billing_company_id.name or '',
                str(rc.cancellation_date) if rc.cancellation_date else '',
                str(rc.travel_date) if rc.travel_date else '',
                rc.booking_executive.name if rc.booking_executive else '',
                rc.employee_code or '',
                rc.document_number or '',
                (rc.passenger_names or '').replace('\n', ', '),
                rc.num_passengers or 0,
                rc.origin_station or '',
                rc.destination_station or '',
                dict(rc._fields['quota'].selection).get(rc.quota, '') if rc.quota else '',
                rc.pnr_number or '',
                rc.train_number or '',
                rc.total_fare or 0.0,
                rc.service_charge or 0.0,
                rc.refund_amount or 0.0,
                dict(rc._fields['mode_of_refund'].selection).get(rc.mode_of_refund, '') if rc.mode_of_refund else '',
                rc.confirmed_by.name if rc.confirmed_by else '',
                dict(rc._fields['state'].selection).get(rc.state, ''),
                rc.remarks or '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        self.write({'excel_file': base64.b64encode(fp.read()), 'file_name': 'train_cancellations.xlsx'})
        fp.close()
        return {'type': 'ir.actions.act_window', 'res_model': self._name, 'res_id': self.id, 'view_mode': 'form', 'target': 'new'}
