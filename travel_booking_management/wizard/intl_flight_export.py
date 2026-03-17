# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, _

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None


class IntlFlightExport(models.TransientModel):
    _name = 'travel.international.flight.booking.export'
    _description = 'Export International Flight Bookings to Excel'

    date_from = fields.Date(string='From Date')
    date_to = fields.Date(string='To Date')
    state_filter = fields.Selection([
        ('all', 'All'), ('draft', 'Draft'), ('confirmed', 'Confirmed'),
        ('done', 'Done'), ('cancelled', 'Cancelled'),
    ], string='Status', default='all')
    excel_file = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='File Name', readonly=True)

    def action_export(self):
        self.ensure_one()
        domain = []
        if self.date_from:
            domain.append(('booking_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('booking_date', '<=', self.date_to))
        if self.state_filter and self.state_filter != 'all':
            domain.append(('state', '=', self.state_filter))
        if self.env.context.get('active_ids') and self.env.context.get('active_model') == 'travel.international.flight.booking':
            domain = [('id', 'in', self.env.context['active_ids'])]
            if self.state_filter and self.state_filter != 'all':
                domain.append(('state', '=', self.state_filter))

        bookings = self.env['travel.international.flight.booking'].search(domain, order='booking_date asc')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Intl Flight Bookings'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = [
            'Booking Ref', 'Billing Company', 'Booking Date', 'Booking Executive',
            'Employee Code', 'Doc No/Req By', 'Trip Type', 'Origin', 'Destination',
            'Return Origin', 'Return Destination', 'Travel Date (Onward)', 'Return Date',
            'Airline', 'Flight No (Onward)', 'Flight No (Return)', 'Class',
            'Passenger(s)', 'No. of Passengers', 'Ticket Number', 'PNR Number',
            'Gross Amount', 'Net Amount', 'Service Charge', 'GST Amount',
            'Mode of Payment', 'Confirmed By', 'Status', 'Remarks',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, bk in enumerate(bookings, 2):
            values = [
                bk.name or '',
                bk.billing_company_id.name or '',
                str(bk.booking_date) if bk.booking_date else '',
                bk.booking_executive.name if bk.booking_executive else '',
                bk.employee_code or '',
                bk.document_number or '',
                dict(bk._fields['trip_type'].selection).get(bk.trip_type, ''),
                bk.origin_city or '',
                bk.destination_city or '',
                bk.return_origin or '',
                bk.return_destination or '',
                str(bk.travel_date_onward) if bk.travel_date_onward else '',
                str(bk.return_date) if bk.return_date else '',
                bk.airline.name if bk.airline else '',
                bk.flight_number or '',
                bk.flight_number_return or '',
                dict(bk._fields['travel_class'].selection).get(bk.travel_class, ''),
                (bk.passenger_names or '').replace('\n', ', '),
                bk.num_passengers or 0,
                bk.ticket_number or '',
                bk.pnr_number or '',
                bk.gross_amount or 0.0,
                bk.total_amount or 0.0,
                bk.service_charge or 0.0,
                bk.gst_amount or 0.0,
                dict(bk._fields['mode_of_payment'].selection).get(bk.mode_of_payment, '') if bk.mode_of_payment else '',
                bk.confirmed_by.name if bk.confirmed_by else '',
                dict(bk._fields['state'].selection).get(bk.state, ''),
                bk.remarks or '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        self.write({'excel_file': base64.b64encode(fp.read()), 'file_name': 'intl_flight_bookings.xlsx'})
        fp.close()
        return {'type': 'ir.actions.act_window', 'res_model': self._name, 'res_id': self.id, 'view_mode': 'form', 'target': 'new'}
