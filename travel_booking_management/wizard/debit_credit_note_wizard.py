# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError

# All booking models that can generate Debit Notes
BOOKING_MODELS = [
    ('travel.domestic.flight.booking', 'Domestic Flight'),
    ('travel.international.flight.booking', 'International Flight'),
    ('travel.hotel.booking', 'Hotel'),
    ('travel.train.booking', 'Train'),
    ('travel.bus.booking', 'Bus'),
    ('travel.car.booking', 'Car'),
    ('travel.insurance.booking', 'Insurance'),
    ('travel.visa.booking', 'Visa'),
    ('travel.package.tour.booking', 'Package Tour'),
    ('travel.event.booking', 'Event'),
]

# All cancellation models that can generate Credit Notes
CANCELLATION_MODELS = [
    ('travel.domestic.flight.cancellation', 'Domestic Flight'),
    ('travel.intl.flight.cancellation', 'International Flight'),
    ('travel.hotel.cancellation', 'Hotel'),
    ('travel.train.cancellation', 'Train'),
    ('travel.bus.cancellation', 'Bus'),
    ('travel.insurance.cancellation', 'Insurance'),
    ('travel.visa.cancellation', 'Visa'),
]


class DebitCreditNoteWizard(models.TransientModel):
    _name = 'travel.debit.credit.note.wizard'
    _description = 'Generate Debit Note / Credit Note'

    note_type = fields.Selection([
        ('debit', 'Debit Note'),
        ('credit', 'Credit Note'),
    ], string='Type', required=True, default='debit')
    partner_id = fields.Many2one(
        'res.partner', string='Billing Company', required=True,
        domain="[('is_company', '=', True)]",
    )
    date_from = fields.Date(string='From Date', required=True)
    date_to = fields.Date(string='To Date', required=True)
    booking_type_filter = fields.Selection([
        ('all', 'All Booking Types'),
        ('domestic_flight', 'Domestic Flight'),
        ('international_flight', 'International Flight'),
        ('hotel', 'Hotel'),
        ('train', 'Train'),
        ('bus', 'Bus'),
        ('car', 'Car'),
        ('insurance', 'Insurance'),
        ('visa', 'Visa'),
        ('package_tour', 'Package Tour'),
        ('event', 'Event'),
    ], string='Booking Type', default='all')
    line_count = fields.Integer(string='Records Found', compute='_compute_preview', store=False)
    total_service_charge = fields.Float(string='Total Service Charge', compute='_compute_preview', store=False)
    total_fare = fields.Float(string='Total Fare', compute='_compute_preview', store=False)

    @api.depends('note_type', 'partner_id', 'date_from', 'date_to', 'booking_type_filter')
    def _compute_preview(self):
        for wiz in self:
            records = wiz._get_matching_records()
            wiz.line_count = len(records)
            wiz.total_service_charge = sum(r.get('service_charge', 0) for r in records)
            wiz.total_fare = sum(r.get('total_fare', 0) for r in records)

    def _get_booking_type_map(self):
        """Map filter keys to model names."""
        return {
            'domestic_flight': 'travel.domestic.flight.booking',
            'international_flight': 'travel.international.flight.booking',
            'hotel': 'travel.hotel.booking',
            'train': 'travel.train.booking',
            'bus': 'travel.bus.booking',
            'car': 'travel.car.booking',
            'insurance': 'travel.insurance.booking',
            'visa': 'travel.visa.booking',
            'package_tour': 'travel.package.tour.booking',
            'event': 'travel.event.booking',
        }

    def _get_cancellation_type_map(self):
        return {
            'domestic_flight': 'travel.domestic.flight.cancellation',
            'international_flight': 'travel.intl.flight.cancellation',
            'hotel': 'travel.hotel.cancellation',
            'train': 'travel.train.cancellation',
            'bus': 'travel.bus.cancellation',
            'insurance': 'travel.insurance.cancellation',
            'visa': 'travel.visa.cancellation',
        }

    def _get_airline_label(self, rec):
        """Get airline display name from selection, many2one, or char field."""
        if hasattr(rec, 'airline') and rec.airline:
            field = rec._fields.get('airline')
            if field and field.type == 'many2one':
                return rec.airline.name or ''
            if field and field.type == 'selection':
                return dict(field.selection).get(rec.airline, rec.airline)
            return rec.airline
        return ''

    def _get_trip_type_label(self, rec):
        if hasattr(rec, 'trip_type') and rec.trip_type:
            return dict(rec._fields['trip_type'].selection).get(rec.trip_type, '')
        if hasattr(rec, 'rental_type') and rec.rental_type:
            return dict(rec._fields['rental_type'].selection).get(rec.rental_type, '')
        return ''

    def _get_class_label(self, rec):
        if hasattr(rec, 'travel_class') and rec.travel_class:
            return dict(rec._fields['travel_class'].selection).get(rec.travel_class, '')
        return ''

    def _extract_record_data(self, rec, model_name, label, is_cancellation=False):
        """Extract a normalized dict of data from any booking/cancellation record."""
        # Passenger names
        pax = ''
        num_pax = 0
        for f in ('passenger_names', 'guest_names'):
            if hasattr(rec, f) and getattr(rec, f):
                pax = getattr(rec, f).strip().replace('\n', ', ')
                num_pax = len([l for l in getattr(rec, f).strip().splitlines() if l.strip()])
                break
        if not pax and hasattr(rec, 'passenger_name') and rec.passenger_name:
            pax = rec.passenger_name
            num_pax = 1

        # Date
        date_val = ''
        if is_cancellation and hasattr(rec, 'cancellation_date'):
            date_val = rec.cancellation_date
        elif hasattr(rec, 'booking_date'):
            date_val = rec.booking_date

        # Origin / Destination
        origin = ''
        destination = ''
        for o_f in ('origin_city', 'origin_station', 'pickup_location', 'location', 'city'):
            if hasattr(rec, o_f) and getattr(rec, o_f):
                origin = getattr(rec, o_f)
                break
        for d_f in ('destination_city', 'destination_station', 'drop_location'):
            if hasattr(rec, d_f) and getattr(rec, d_f):
                destination = getattr(rec, d_f)
                break

        # Travel date / Return date
        travel_date = ''
        return_date = ''
        for tf in ('travel_date_onward', 'travel_date', 'checkin_date', 'pickup_date'):
            if hasattr(rec, tf) and getattr(rec, tf):
                travel_date = getattr(rec, tf)
                break
        for rf in ('return_date', 'checkout_date', 'drop_date'):
            if hasattr(rec, rf) and getattr(rec, rf):
                return_date = getattr(rec, rf)
                break

        # PNR / Ticket
        pnr = ''
        ticket = ''
        for pf in ('pnr_number', 'reference_number', 'booking_reference'):
            if hasattr(rec, pf) and getattr(rec, pf):
                pnr = getattr(rec, pf)
                break
        if hasattr(rec, 'ticket_number') and rec.ticket_number:
            ticket = rec.ticket_number

        # Flight number
        flight_no = ''
        if hasattr(rec, 'flight_number') and rec.flight_number:
            flight_no = rec.flight_number

        # Fare / amounts
        total_fare = 0.0
        for af in ('total_amount', 'fare', 'total_fare', 'gross_amount', 'amount'):
            if hasattr(rec, af) and getattr(rec, af):
                total_fare = getattr(rec, af)
                break

        service_charge = 0.0
        if hasattr(rec, 'service_charge') and rec.service_charge:
            service_charge = rec.service_charge

        refund_amount = 0.0
        if hasattr(rec, 'refund_amount') and rec.refund_amount:
            refund_amount = rec.refund_amount

        cancel_cost = max(total_fare - refund_amount, 0) if is_cancellation and refund_amount else 0.0

        return {
            'record': rec,
            'model': model_name,
            'booking_type_label': label,
            'name': rec.name,
            'date': date_val,
            'num_pax': num_pax,
            'passenger_names': pax,
            'employee_code': rec.employee_code if hasattr(rec, 'employee_code') else '',
            'document_number': rec.document_number if hasattr(rec, 'document_number') else '',
            'origin': origin,
            'destination': destination,
            'trip_type': self._get_trip_type_label(rec),
            'travel_date': travel_date,
            'return_date': return_date,
            'pnr_ticket': ticket or pnr,
            'pnr_number': pnr,
            'flight_number': flight_no,
            'travel_class': self._get_class_label(rec),
            'airline': self._get_airline_label(rec),
            'total_fare': total_fare,
            'service_charge': service_charge,
            'refund_amount': refund_amount,
            'cancel_cost': cancel_cost,
        }

    def _get_matching_records(self):
        """Find all booking or cancellation records matching the wizard filters."""
        self.ensure_one()
        if not self.partner_id or not self.date_from or not self.date_to:
            return []

        results = []
        if self.note_type == 'debit':
            type_map = self._get_booking_type_map()
            models_to_search = (
                {self.booking_type_filter: type_map[self.booking_type_filter]}
                if self.booking_type_filter != 'all'
                else type_map
            )
            for key, model_name in models_to_search.items():
                Model = self.env[model_name]
                label = dict(self._fields['booking_type_filter'].selection).get(key, key)
                # Determine partner field
                partner_field = 'partner_id' if model_name == 'travel.event.booking' else 'billing_company_id'
                domain = [
                    (partner_field, '=', self.partner_id.id),
                    ('booking_date', '>=', self.date_from),
                    ('booking_date', '<=', self.date_to),
                    ('state', '=', 'confirmed'),
                ]
                for rec in Model.search(domain):
                    data = self._extract_record_data(rec, model_name, label)
                    if data['service_charge'] > 0 or data['total_fare'] > 0:
                        results.append(data)
        else:  # credit
            type_map = self._get_cancellation_type_map()
            models_to_search = (
                {self.booking_type_filter: type_map[self.booking_type_filter]}
                if self.booking_type_filter != 'all' and self.booking_type_filter in type_map
                else type_map
            )
            for key, model_name in models_to_search.items():
                Model = self.env[model_name]
                label = dict(self._fields['booking_type_filter'].selection).get(key, key)
                domain = [
                    ('billing_company_id', '=', self.partner_id.id),
                    ('cancellation_date', '>=', self.date_from),
                    ('cancellation_date', '<=', self.date_to),
                    ('state', 'in', ['confirmed', 'refunded']),
                ]
                for rec in Model.search(domain):
                    data = self._extract_record_data(rec, model_name, label, is_cancellation=True)
                    if data['service_charge'] > 0 or data['total_fare'] > 0:
                        results.append(data)
        return results

    def action_preview(self):
        """Refresh the wizard."""
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_generate(self):
        """Generate TWO separate documents:
        1. Tax Invoice — service charge lines with CGST+SGST
        2. Debit Note — fare reimbursement lines with NO tax
        For credit notes: Credit Note (service refund) + Credit Note (fare refund).
        """
        self.ensure_one()
        records = self._get_matching_records()
        if not records:
            raise UserError(_("No matching records found for the selected criteria."))

        company = self.env.company
        cgst = self.env['account.tax'].search([
            ('name', '=', 'CGST 9%'), ('company_id', '=', company.id),
            ('type_tax_use', '=', 'sale'),
        ], limit=1)
        sgst = self.env['account.tax'].search([
            ('name', '=', 'SGST 9%'), ('company_id', '=', company.id),
            ('type_tax_use', '=', 'sale'),
        ], limit=1)
        tax_ids = [t.id for t in (cgst | sgst) if t]

        sc_product = self.env['product.product'].search(
            [('default_code', '=', 'product_service_charge')], limit=1)
        if not sc_product:
            raise UserError(_("Travel Service Charge product not found."))

        fare_product = self.env['product.product'].search(
            [('default_code', '=', 'product_travel_fare')], limit=1)
        if not fare_product:
            raise UserError(_("Travel Fare Reimbursement product not found."))

        is_credit = self.note_type == 'credit'

        # Collect SO/cancellation references
        so_names = []
        so_ids = set()
        cancel_names = []
        for data in records:
            rec = data['record']
            cancel_names.append(data['name'])
            if hasattr(rec, 'sale_order_id') and rec.sale_order_id:
                so_names.append(rec.sale_order_id.name)
                so_ids.add(rec.sale_order_id.id)
        origin_str = ', '.join(sorted(set(so_names))) if so_names else (
            ', '.join(cancel_names) if cancel_names else False)

        created_moves = self.env['account.move']

        # ── 1. TAX INVOICE / CREDIT NOTE TAX INVOICE (Service Charge + GST) ──
        # For bookings: out_invoice (Tax Invoice)
        # For cancellations: out_invoice (Tax Invoice with CN labels — cancellation
        #   service charge is a NEW charge, not a refund)
        sc_records = [r for r in records if r['service_charge'] > 0]
        tax_invoice = None
        if sc_records:
            sc_lines = []
            for data in sc_records:
                desc = f"Service Charges for {data['booking_type_label']} Booking"
                line_vals = {
                    'product_id': sc_product.id,
                    'name': desc,
                    'quantity': 1,
                    'price_unit': data['service_charge'],
                    'tax_ids': [(6, 0, tax_ids)],
                }
                rec = data['record']
                if hasattr(rec, 'sale_order_id') and rec.sale_order_id:
                    sc_sol = rec.sale_order_id.order_line.filtered(
                        lambda l: l.product_id.default_code == 'product_service_charge'
                    )
                    if sc_sol:
                        line_vals['sale_line_ids'] = [(4, sc_sol[0].id)]
                sc_lines.append((0, 0, line_vals))

            # Always out_invoice — service charge is always billed TO the customer
            sc_label = 'Credit Note Tax Invoice' if is_credit else 'Tax Invoice'
            tax_invoice = self.env['account.move'].create({
                'move_type': 'out_invoice',
                'partner_id': self.partner_id.id,
                'currency_id': company.currency_id.id,
                'invoice_date': fields.Date.today(),
                'invoice_origin': origin_str,
                'narration': f"{sc_label} for {self.partner_id.name} ({self.date_from} to {self.date_to})",
                'is_cancellation_service_invoice': is_credit,
                'invoice_line_ids': sc_lines,
            })
            # Attach service charge annexure
            self._attach_service_charge_annexure(tax_invoice, sc_records)
            created_moves |= tax_invoice

        # ── 2. DEBIT NOTE / CREDIT NOTE (Fare) ──
        fare_records = [r for r in records if r['total_fare'] > 0]
        debit_note = None
        if fare_records:
            fare_lines = []
            for data in fare_records:
                btype = data['booking_type_label']
                desc_map = {
                    'Domestic Flight': 'Dom. Air Fare Reimbursement As Per Attached Annex.',
                    'International Flight': 'Intl. Air Fare Reimbursement As Per Attached Annex.',
                    'Hotel': 'Hotel Fare Reimbursement As Per Attached Annex.',
                    'Train': 'Train Fare Reimbursement As Per Attached Annex.',
                    'Bus': 'Bus Fare Reimbursement As Per Attached Annex.',
                    'Car': 'Car Fare Reimbursement As Per Attached Annex.',
                    'Insurance': 'Insurance Fare Reimbursement As Per Attached Annex.',
                    'Visa': 'Visa Fare Reimbursement As Per Attached Annex.',
                    'Package Tour': 'Package Tour Fare Reimbursement As Per Attached Annex.',
                    'Event': 'Event Fare Reimbursement As Per Attached Annex.',
                }
                desc = desc_map.get(btype, f'{btype} Fare Reimbursement As Per Attached Annex.')
                line_vals = {
                    'product_id': fare_product.id,
                    'name': desc,
                    'quantity': 1,
                    'price_unit': data['total_fare'],
                    'tax_ids': [(6, 0, [])],  # NO tax on fare reimbursement
                }
                rec = data['record']
                if hasattr(rec, 'sale_order_id') and rec.sale_order_id:
                    fare_sol = rec.sale_order_id.order_line.filtered(
                        lambda l: l.product_id.default_code == 'product_travel_fare'
                    )
                    if fare_sol:
                        line_vals['sale_line_ids'] = [(4, fare_sol[0].id)]
                fare_lines.append((0, 0, line_vals))

            # For bookings: out_invoice (Debit Note)
            # For cancellations: out_refund (Credit Note for fare refund)
            fare_move_type = 'out_refund' if is_credit else 'out_invoice'
            dn_label = 'Credit Note (Fare)' if is_credit else 'Debit Note'
            debit_note = self.env['account.move'].create({
                'move_type': fare_move_type,
                'partner_id': self.partner_id.id,
                'currency_id': company.currency_id.id,
                'invoice_date': fields.Date.today(),
                'invoice_origin': origin_str,
                'narration': f"{dn_label} for {self.partner_id.name} ({self.date_from} to {self.date_to})",
                'invoice_line_ids': fare_lines,
            })
            # Attach fare charge annexure
            self._attach_fare_charge_annexure(debit_note, fare_records)
            created_moves |= debit_note

        if not created_moves:
            raise UserError(_("No service charge or fare lines found to generate documents."))

        if len(created_moves) == 1:
            label = 'Tax Invoice' if tax_invoice and not debit_note else ('Debit Note' if not is_credit else 'Credit Note')
            return {
                'type': 'ir.actions.act_window',
                'name': _(label),
                'res_model': 'account.move',
                'res_id': created_moves[0].id,
                'view_mode': 'form',
                'target': 'current',
            }
        else:
            return {
                'type': 'ir.actions.act_window',
                'name': _('Generated Documents'),
                'res_model': 'account.move',
                'domain': [('id', 'in', created_moves.ids)],
                'view_mode': 'list,form',
                'target': 'current',
            }

    def _get_excel_styles(self):
        """Return common Excel styles for annexures."""
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return None
        return {
            'thin_border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            ),
            'header_font': Font(bold=True, color='FFFFFF', size=9),
            'header_fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
            'header_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'bold_font': Font(bold=True, size=10),
        }

    def _attach_service_charge_annexure(self, invoice, sc_records):
        """Generate Service Charge Excel annexure and attach to Tax Invoice."""
        import base64
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return

        styles = self._get_excel_styles()
        if not styles:
            return
        is_credit = self.note_type == 'credit'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Service Charge Annexure'

        # For cancellations: "ANNEXURE-1" + "DETAILS OF X BOOKING" (matching reference)
        # For bookings: "ANNEXURE - SERVICE CHARGE"
        first_label = sc_records[0]['booking_type_label'] if sc_records else 'Travel'
        if is_credit:
            title = 'ANNEXURE-1'
            self._write_annexure_header(ws, title,
                                        'Credit Note',
                                        styles['bold_font'])
            ws.cell(row=4, column=1,
                    value=f'DETAILS OF {first_label.upper()} BOOKING').font = Font(bold=True, size=11)
            # Partner and reference info
            ws.cell(row=5, column=1, value='To,').font = Font(size=9)
            ws.cell(row=5, column=10, value=str(self.date_from) if self.date_from else '').font = Font(size=9)
            ws.cell(row=6, column=1, value=self.partner_id.name or '').font = Font(bold=True, size=9)
            inv_ref = invoice.name or f'Draft-{invoice.id}'
            ws.cell(row=5, column=14, value=f'DCN-{inv_ref}').font = Font(bold=True, size=9)
            ws.cell(row=6, column=14, value=f'Ref: CN/DCN-{inv_ref}').font = Font(size=9)
            data_start_row = 8
        else:
            self._write_annexure_header(ws, 'ANNEXURE - SERVICE CHARGE',
                                        'Tax Invoice',
                                        styles['bold_font'])
            data_start_row = 6

        row = data_start_row
        sc_headers = [
            'SR No.', 'Cancel Date' if is_credit else 'Booking Date',
            'No of', 'Passenger(s) Name', 'Emp. Code', 'Doc. No.',
            'From Origin', 'To Destination', 'Travel Type',
            'Travel Date', 'Return Date', 'PNR / Ticket No.', 'PNR No.',
            'Flight No.', 'Class of Booking', 'Airline Name',
            'Service Charge', 'CGST 9%', 'SGST 9%', 'IGST', 'Total ₹',
            'Inv No.', 'Remark',
        ]
        for col, h in enumerate(sc_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_align']
            cell.border = styles['thin_border']

        total_sc = total_cgst = total_sgst = total_total = 0
        for idx, data in enumerate(sc_records, 1):
            row += 1
            sc = data['service_charge']
            cgst_amt = round(sc * 0.09, 2)
            sgst_amt = round(sc * 0.09, 2)
            line_total = round(sc + cgst_amt + sgst_amt, 2)
            total_sc += sc
            total_cgst += cgst_amt
            total_sgst += sgst_amt
            total_total += line_total
            vals = [
                idx, str(data['date']) if data['date'] else '',
                data['num_pax'], data['passenger_names'],
                data['employee_code'] or '', data['document_number'] or '',
                data['origin'], data['destination'], data['trip_type'],
                str(data['travel_date']) if data['travel_date'] else '',
                str(data['return_date']) if data['return_date'] else '',
                data['pnr_ticket'], data['pnr_number'],
                data['flight_number'], data['travel_class'], data['airline'],
                sc, cgst_amt, sgst_amt, 0, line_total,
                invoice.name or f'Draft-{invoice.id}', '',
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = styles['thin_border']

        row += 1
        ws.cell(row=row, column=3, value='Grand Total').font = styles['bold_font']
        for col, val in [(17, total_sc), (18, total_cgst), (19, total_sgst), (20, 0), (21, total_total)]:
            cell = ws.cell(row=row, column=col, value=round(val, 2))
            cell.font = styles['bold_font']
            cell.border = styles['thin_border']

        self._auto_width(ws)
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        self.env['ir.attachment'].create({
            'name': f"Annexure_ServiceCharge_{invoice.name or 'draft'}.xlsx",
            'type': 'binary',
            'datas': base64.b64encode(fp.read()),
            'res_model': 'account.move',
            'res_id': invoice.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        fp.close()

    def _attach_fare_charge_annexure(self, invoice, fare_records):
        """Generate Fare Charge Excel annexure and attach to Debit Note."""
        import base64
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return

        styles = self._get_excel_styles()
        if not styles:
            return
        is_credit = self.note_type == 'credit'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fare Charge Annexure'
        self._write_annexure_header(ws, 'ANNEXURE-1',
                                    'Credit Note' if is_credit else 'Debit Note',
                                    styles['bold_font'])
        first_label = fare_records[0]['booking_type_label'] if fare_records else 'Travel'
        ws.cell(row=4, column=1, value=f'DETAILS OF {first_label.upper()} BOOKING').font = Font(bold=True, size=11)
        # Partner and reference info
        ws.cell(row=5, column=1, value='To,').font = Font(size=9)
        ws.cell(row=5, column=10, value=str(self.date_from) if self.date_from else '').font = Font(size=9)
        ws.cell(row=6, column=1, value=self.partner_id.name or '').font = Font(bold=True, size=9)
        inv_ref = invoice.name or f'Draft-{invoice.id}'
        prefix = 'CN' if is_credit else 'DN'
        ws.cell(row=5, column=14, value=f'{prefix}/{inv_ref}').font = Font(bold=True, size=9)
        ws.cell(row=6, column=14, value=f'Ref: {invoice.invoice_origin or inv_ref}').font = Font(size=9)

        row = 8
        fare_headers = [
            'SR No.', 'Cancel Date' if is_credit else 'Booking Date',
            'No of Pax.', 'Passenger(s) Name', 'Emp. Code', 'Doc. No.',
            'From Origin', 'To Destination', 'Travel Type',
            'Travel Date', 'Return Date', 'PNR / Ticket No.', 'PNR No.',
            'Flight No.', 'Flight No. Return', 'Class of Booking', 'Airline Name',
            'Basic Amount', 'GST Amount', 'Net Amount',
            'Remark',
        ]
        for col, h in enumerate(fare_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_align']
            cell.border = styles['thin_border']

        total_basic = total_gst = total_net = 0
        for idx, data in enumerate(fare_records, 1):
            row += 1
            rec = data['record']
            net_amount = data['total_fare']
            gst_amt = 0.0
            if hasattr(rec, 'gst_amount') and rec.gst_amount:
                gst_amt = rec.gst_amount
            basic_amt = net_amount - gst_amt if gst_amt else net_amount
            total_basic += basic_amt
            total_gst += gst_amt
            total_net += net_amount
            flight_return = ''
            if hasattr(rec, 'flight_number_return') and rec.flight_number_return:
                flight_return = rec.flight_number_return
            vals = [
                idx, str(data['date']) if data['date'] else '',
                data['num_pax'], data['passenger_names'],
                data['employee_code'] or '', data['document_number'] or '',
                data['origin'], data['destination'], data['trip_type'],
                str(data['travel_date']) if data['travel_date'] else '',
                str(data['return_date']) if data['return_date'] else '',
                data['pnr_ticket'], data['pnr_number'],
                data['flight_number'], flight_return,
                data['travel_class'], data['airline'],
                basic_amt, gst_amt, net_amount,
                '',
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = styles['thin_border']

        row += 1
        ws.cell(row=row, column=3, value='Grand Total').font = styles['bold_font']
        for col, val in [(18, total_basic), (19, total_gst), (20, total_net)]:
            cell = ws.cell(row=row, column=col, value=round(val, 2))
            cell.font = styles['bold_font']
            cell.border = styles['thin_border']

        self._auto_width(ws)
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        self.env['ir.attachment'].create({
            'name': f"Annexure_FareCharge_{invoice.name or 'draft'}.xlsx",
            'type': 'binary',
            'datas': base64.b64encode(fp.read()),
            'res_model': 'account.move',
            'res_id': invoice.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        fp.close()

    def _write_annexure_header(self, ws, title, note_label, bold_font):
        """Write the company header rows on the annexure sheet."""
        from openpyxl.styles import Font
        company = self.env.company
        ws.cell(row=1, column=1, value='THE EARTH').font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=company.street or '').font = Font(size=9)
        ws.cell(row=3, column=1, value=f'GSTIN: {company.vat or ""}').font = Font(size=9)
        ws.cell(row=4, column=1, value=title).font = Font(bold=True, size=12)
        ws.cell(row=4, column=10, value=f'To: {self.partner_id.name}').font = bold_font
        ws.cell(row=5, column=10, value=f'{self.date_from} to {self.date_to}').font = Font(size=9)

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
