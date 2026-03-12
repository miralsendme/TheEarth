# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, _

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None


class HotelCancellationExport(models.TransientModel):
    _name = 'travel.hotel.cancellation.export'
    _description = 'Export Hotel Cancellations to Excel'

    date_from = fields.Date(string='From Date')
    date_to = fields.Date(string='To Date')
    state_filter = fields.Selection([
        ('all', 'All'), ('draft', 'Draft'), ('confirmed', 'Confirmed'),
        ('refunded', 'Refunded'), ('rejected', 'Rejected'),
    ], string='Status', default='all')
    excel_file = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='File Name', readonly=True)

    def action_export(self):
        self.ensure_one()
        domain = []
        if self.date_from:
            domain.append(('cancellation_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('cancellation_date', '<=', self.date_to))
        if self.state_filter and self.state_filter != 'all':
            domain.append(('state', '=', self.state_filter))
        if self.env.context.get('active_ids') and self.env.context.get('active_model') == 'travel.hotel.cancellation':
            domain = [('id', 'in', self.env.context['active_ids'])]
            if self.state_filter and self.state_filter != 'all':
                domain.append(('state', '=', self.state_filter))

        records = self.env['travel.hotel.cancellation'].search(domain, order='cancellation_date asc')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Hotel Cancellations'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = [
            'Cancellation Ref', 'Billing Company', 'Cancellation Date', 'Booking Executive',
            'Employee Code', 'Doc No/Req By', 'Guest(s)', 'No. of Guests Cancelled',
            'Location', 'Hotel Name', 'Check-in', 'Check-out', 'Nights Cancelled',
            'Booking Reference', 'Total Amount', 'Refund Amount',
            'Mode of Refund', 'Confirmed By', 'Status', 'Remarks',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, rc in enumerate(records, 2):
            values = [
                rc.name or '',
                rc.billing_company_id.name or '',
                str(rc.cancellation_date) if rc.cancellation_date else '',
                rc.booking_executive.name if rc.booking_executive else '',
                rc.employee_code or '',
                rc.document_number or '',
                (rc.guest_names or '').replace('\n', ', '),
                rc.num_cancelled_guests or 0,
                rc.location or '',
                rc.hotel_name or '',
                str(rc.checkin_date) if rc.checkin_date else '',
                str(rc.checkout_date) if rc.checkout_date else '',
                rc.num_nights_cancelled or 0,
                rc.booking_reference or '',
                rc.total_amount or 0.0,
                rc.refund_amount or 0.0,
                dict(rc._fields['mode_of_refund'].selection).get(rc.mode_of_refund, '') if rc.mode_of_refund else '',
                rc.confirmed_by.name if rc.confirmed_by else '',
                dict(rc._fields['state'].selection).get(rc.state, ''),
                rc.remarks or '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        self.write({'excel_file': base64.b64encode(fp.read()), 'file_name': 'hotel_cancellations.xlsx'})
        fp.close()
        return {'type': 'ir.actions.act_window', 'res_model': self._name, 'res_id': self.id, 'view_mode': 'form', 'target': 'new'}
