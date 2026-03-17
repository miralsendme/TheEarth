# -*- coding: utf-8 -*-
import base64
from datetime import date as _date
from odoo import models, fields, api, _
from odoo.exceptions import UserError

# Cancellation type → prefix mapping
CANCELLATION_PREFIX_MAP = {
    'train': 'CT',
    'bus': 'CB',
    'domestic_flight': 'CDF',
    'international_flight': 'CIF',
    'hotel': 'CH',
    'insurance': 'CTI',
    'visa': 'CV',
    'car': 'CC',
    'package_tour': 'CPT',
    'event': 'CE',
}


def get_cancellation_ref(env, cancellation_type):
    """Generate cancellation reference like CTSO/1/25-26.
    Uses shared sequence 'travel.cancellation.ref' with Indian FY date ranges.
    """
    prefix = CANCELLATION_PREFIX_MAP.get(cancellation_type, 'C')
    today = _date.today()
    if today.month >= 4:
        fy_start = _date(today.year, 4, 1)
        fy_end = _date(today.year + 1, 3, 31)
        fy_str = f"{today.year % 100}-{(today.year + 1) % 100:02d}"
    else:
        fy_start = _date(today.year - 1, 4, 1)
        fy_end = _date(today.year, 3, 31)
        fy_str = f"{(today.year - 1) % 100}-{today.year % 100:02d}"

    seq = env['ir.sequence'].search([('code', '=', 'travel.cancellation.ref')], limit=1)
    if not seq:
        return False

    date_range = env['ir.sequence.date_range'].search([
        ('sequence_id', '=', seq.id),
        ('date_from', '=', fy_start),
        ('date_to', '=', fy_end),
    ], limit=1)
    if not date_range:
        date_range = env['ir.sequence.date_range'].create({
            'sequence_id': seq.id,
            'date_from': fy_start,
            'date_to': fy_end,
            'number_next': 1,
        })

    number = str(date_range._next())
    return f"{prefix}SO/{number}/{fy_str}"

# Mapping of mode_of_payment keys to vendor names for Purchase Orders
VENDOR_MAP = {
    'air_asia': 'Air Asia (India) Limited',
    'akasa_airline': 'Akasa Airline',
    'akbar_offshore': 'AKBAR OFFSHORE PVT LTD',
    'akbar_new': 'Akbar Online Booking Company Pvt Ltd',
    'akbar_old': 'Akbar Online Booking Company Pvt Ltd',
    'aman_travels': 'Aman Travels Ltd',
    'interglobe': 'Interglobe Aviation Limited',
    'mmt_wallet': 'MakeMyTrip',
    'pcc_akbar_offshore': 'PCC AKBAR OFFSHORE PVT LTD',
    'plus_wallet': 'Plus Wallet',
    'riya_offline': 'Riya Travels & Tours',
    'riya_online': 'Riya Travel & Tours',
    'spicejet': 'Spicejet Limited',
    # Credit card payments — vendor is The Earth itself (internal)
    'axis_cc_vistara_deep_1236': False,
    'axis_debit_deep_2100': False,
    'hdfc_cc_deep_0943': False,
    'hdfc_cc_deep_6223': False,
    'ketan_axis_cc_ace_7929': False,
    'axis_cc_nirav_2281': False,
    'axis_debit_nirav_2448': False,
    'hdfc_cc_nirav_9912': False,
    'hdfc_nirav_cc_1583': False,
    'hdfc_nirav_cc_7122': False,
    'hdfc_earth_cc_3026': False,
    'hdfc_earth_cc_7209': False,
    'hdfc_earth_card_0481': False,
    'earth_hdfc_cc_1554': False,
    'icici_deep_cc_0003': False,
    'icici_ketan_cc_9005': False,
}

# Hotel-specific service provider to vendor mapping
HOTEL_PROVIDER_MAP = {
    'aman_travels': 'Aman Travels Ltd',
    'makemytrip': 'MakeMyTrip',
    'treebo': 'Treebo',
    'goibibo': 'Goibibo',
    'airbnb': 'Airbnb',
    'meril_travel_desk': 'Meril Travel Desk',
    'travel_plus': 'Travel Plus',
    'other': False,
}


class TravelSalePurchaseMixin(models.AbstractModel):
    _name = 'travel.sale.purchase.mixin'
    _description = 'Mixin for auto-generating Draft SO and PO from bookings'

    sale_order_id = fields.Many2one('sale.order', string='Sales Order', readonly=True, copy=False)
    purchase_order_id = fields.Many2one('purchase.order', string='Purchase Order', readonly=True, copy=False)

    @api.model
    def _name_search(self, name='', domain=None, operator='ilike', limit=None, order=None):
        """Allow searching bookings by their linked SO name (e.g. DFSO/5/25-26)."""
        res = super()._name_search(name=name, domain=domain, operator=operator, limit=limit, order=order)
        if name and operator in ('ilike', 'like', '=ilike', '=like', '='):
            so_bookings = self.search([
                ('sale_order_id.name', operator, name),
            ], limit=limit, order=order)
            if so_bookings:
                existing_ids = {r if isinstance(r, int) else r[0] for r in res}
                for rec in so_bookings:
                    if rec.id not in existing_ids:
                        res.append(rec.id)
        return res

    @api.depends('name', 'sale_order_id', 'sale_order_id.name')
    def _compute_display_name(self):
        """Show SO reference alongside booking name for easier identification."""
        for rec in self:
            name = rec.name or ''
            if rec.sale_order_id and rec.sale_order_id.name:
                name = f"{name} [{rec.sale_order_id.name}]"
            rec.display_name = name

    # Mapping from booking type label → SO prefix
    SO_PREFIX_MAP = {
        'Car': 'C',
        'Domestic Flight': 'DF',
        'International Flight': 'IF',
        'Train': 'T',
        'Bus': 'B',
        'Hotel': 'H',
        'Event': 'E',
        'Visa': 'V',
        'Insurance': 'TI',
        'Package Tour': 'PT',
    }

    # Mapping from booking type label → per-type sequence code
    SO_SEQ_CODE_MAP = {
        'Car': 'travel.so.car',
        'Domestic Flight': 'travel.so.domestic.flight',
        'International Flight': 'travel.so.international.flight',
        'Train': 'travel.so.train',
        'Bus': 'travel.so.bus',
        'Hotel': 'travel.so.hotel',
        'Event': 'travel.so.event',
        'Visa': 'travel.so.visa',
        'Insurance': 'travel.so.insurance',
        'Package Tour': 'travel.so.package.tour',
    }

    def _get_booking_description(self):
        """Override in each booking model to return a descriptive line name."""
        return self.name or 'Travel Booking'

    def _get_booking_type_label(self):
        """Override in each booking model to return the booking type label."""
        return 'Travel'

    def _get_so_prefix(self):
        """Return the SO prefix for this booking type (e.g. 'C' for Car)."""
        return self.SO_PREFIX_MAP.get(self._get_booking_type_label(), '')

    @staticmethod
    def _get_indian_fy_string(dt=None):
        """Return Indian financial year string like '25-26' for a given date.
        Indian FY runs April 1 to March 31.
        If month >= April, FY = current_year - next_year, else FY = prev_year - current_year.
        """
        from datetime import date as dt_date
        if dt is None:
            dt = dt_date.today()
        if dt.month >= 4:
            return f"{dt.year % 100}-{(dt.year + 1) % 100:02d}"
        else:
            return f"{(dt.year - 1) % 100}-{dt.year % 100:02d}"

    @api.model
    def _generate_booking_ref(self, booking_type_label):
        """Generate booking reference in same format as SO: {prefix}SO/{number}/{FY}.
        Uses a per-type sequence so each booking type has independent numbering.
        """
        prefix = self.SO_PREFIX_MAP.get(booking_type_label, '')
        seq_code = self.SO_SEQ_CODE_MAP.get(booking_type_label)
        if not prefix or not seq_code:
            return False
        fy = self._get_indian_fy_string()
        number = self._get_next_number_for_seq(seq_code)
        return f"{prefix}SO/{number}/{fy}"

    def _get_next_number_for_seq(self, seq_code):
        """Get the next sequence number from a specific sequence code,
        ensuring the date range is aligned to Indian FY (April 1 - March 31).
        """
        from datetime import date as dt_date
        today = dt_date.today()
        if today.month >= 4:
            fy_start = dt_date(today.year, 4, 1)
            fy_end = dt_date(today.year + 1, 3, 31)
        else:
            fy_start = dt_date(today.year - 1, 4, 1)
            fy_end = dt_date(today.year, 3, 31)

        seq = self.env['ir.sequence'].sudo().search([('code', '=', seq_code)], limit=1)
        if not seq:
            return '1'

        date_range = self.env['ir.sequence.date_range'].sudo().search([
            ('sequence_id', '=', seq.id),
            ('date_from', '=', fy_start),
            ('date_to', '=', fy_end),
        ], limit=1)
        if not date_range:
            date_range = self.env['ir.sequence.date_range'].sudo().create({
                'sequence_id': seq.id,
                'date_from': fy_start,
                'date_to': fy_end,
                'number_next': 1,
            })

        return str(date_range.sudo()._next())

    def _get_ticket_amount(self):
        """Return the ticket/fare amount (non-taxable). Override if field name differs."""
        if hasattr(self, 'total_amount') and self.total_amount:
            return self.total_amount
        if hasattr(self, 'amount') and self.amount:
            return self.amount
        return 0.0

    def _get_service_charge_amount(self):
        """Return the service charge amount (taxable). Override if not applicable."""
        if hasattr(self, 'service_charge') and self.service_charge:
            return self.service_charge
        return 0.0

    def _get_customer_partner(self):
        """Return the customer (billing company) for the SO."""
        if hasattr(self, 'billing_company_id') and self.billing_company_id:
            return self.billing_company_id
        return False

    def _get_vendor_partner(self):
        """Return or create the vendor partner for the PO."""
        vendor_name = False

        # For hotel bookings, prefer booking_service_provider
        if hasattr(self, 'booking_service_provider') and self.booking_service_provider:
            vendor_name = HOTEL_PROVIDER_MAP.get(self.booking_service_provider)

        # Fall back to mode_of_payment
        if not vendor_name and hasattr(self, 'mode_of_payment') and self.mode_of_payment:
            vendor_name = VENDOR_MAP.get(self.mode_of_payment)

        if not vendor_name:
            return False

        # Find or create vendor partner
        partner = self.env['res.partner'].search([
            ('name', '=ilike', vendor_name),
            ('supplier_rank', '>', 0),
        ], limit=1)
        if not partner:
            partner = self.env['res.partner'].search([
                ('name', '=ilike', vendor_name),
            ], limit=1)
        if not partner:
            partner = self.env['res.partner'].create({
                'name': vendor_name,
                'supplier_rank': 1,
                'is_company': True,
            })
        elif not partner.supplier_rank:
            partner.sudo().write({'supplier_rank': partner.supplier_rank + 1})
        return partner

    def _get_passenger_info(self):
        """Return passenger/guest names for SO line description."""
        for field in ('passenger_names', 'guest_names', 'passenger_name'):
            if hasattr(self, field):
                val = getattr(self, field)
                if val:
                    return val.strip()
        return ''

    @staticmethod
    def _strip_name_prefix(name):
        """Strip MR/MRS/MS/DR prefix from a passenger name for employee lookup."""
        import re
        return re.sub(r'^(MR\s+|MRS\s+|MS\s+|DR\s+)', '', name.strip(), flags=re.IGNORECASE).strip()

    def _get_passenger_list(self):
        """Return list of passenger names for annexure."""
        for field in ('passenger_names', 'guest_names'):
            if hasattr(self, field):
                val = getattr(self, field)
                if val:
                    return [n.strip() for n in val.strip().splitlines() if n.strip()]
        if hasattr(self, 'passenger_name') and self.passenger_name:
            return [self.passenger_name.strip()]
        return []

    def _get_pnr_info(self):
        """Return PNR/ticket number if available."""
        for field in ('pnr_number', 'ticket_number', 'reference_number'):
            if hasattr(self, field):
                val = getattr(self, field)
                if val:
                    return val.strip()
        return ''

    def _get_travel_dates_info(self):
        """Return a list of dicts with travel date info for annexure. Override per model."""
        dates = []
        # Flight-style: onward + return
        if hasattr(self, 'travel_date_onward') and self.travel_date_onward:
            dates.append({'label': 'Onward', 'date': self.travel_date_onward})
            if hasattr(self, 'return_date') and self.return_date:
                dates.append({'label': 'Return', 'date': self.return_date})
        # Train/Bus: single travel_date
        elif hasattr(self, 'travel_date') and self.travel_date:
            dates.append({'label': 'Travel Date', 'date': self.travel_date})
        # Hotel/Event: checkin/checkout
        elif hasattr(self, 'checkin_date') and self.checkin_date:
            dates.append({'label': 'Check-in', 'date': self.checkin_date})
            if hasattr(self, 'checkout_date') and self.checkout_date:
                dates.append({'label': 'Check-out', 'date': self.checkout_date})
        # Car: pickup/drop
        elif hasattr(self, 'pickup_date') and self.pickup_date:
            dates.append({'label': 'Pickup', 'date': self.pickup_date})
            if hasattr(self, 'drop_date') and self.drop_date:
                dates.append({'label': 'Drop', 'date': self.drop_date})
        # Booking date as fallback
        if not dates and hasattr(self, 'booking_date') and self.booking_date:
            dates.append({'label': 'Booking Date', 'date': self.booking_date})
        return dates

    def _get_route_info(self):
        """Return origin-destination info if available."""
        if hasattr(self, 'origin_city') and self.origin_city:
            route = self.origin_city
            if hasattr(self, 'destination_city') and self.destination_city:
                route += ' → ' + self.destination_city
            return route
        if hasattr(self, 'origin_station') and self.origin_station:
            route = self.origin_station.display_name if hasattr(self.origin_station, 'display_name') else str(self.origin_station)
            if hasattr(self, 'destination_station') and self.destination_station:
                dest = self.destination_station.display_name if hasattr(self.destination_station, 'display_name') else str(self.destination_station)
                route += ' → ' + dest
            return route
        if hasattr(self, 'pickup_location') and self.pickup_location:
            route = self.pickup_location
            if hasattr(self, 'drop_location') and self.drop_location:
                route += ' → ' + self.drop_location
            return route
        if hasattr(self, 'location') and self.location:
            return self.location
        return ''

    def _get_or_create_product(self, xmlid, name, product_type='service'):
        """Get or create a product by xmlid for SO/PO lines."""
        product = self.env.ref(f'travel_booking_management.{xmlid}', raise_if_not_found=False)
        if not product:
            product = self.env['product.product'].search([('default_code', '=', xmlid)], limit=1)
        if not product:
            product = self.env['product.product'].create({
                'name': name,
                'default_code': xmlid,
                'type': product_type,
                'sale_ok': True,
                'purchase_ok': True,
                'taxes_id': [(5, 0, 0)],
                'supplier_taxes_id': [(5, 0, 0)],
            })
        return product

    def _get_fare_product(self):
        return self._get_or_create_product('product_travel_fare', 'Travel Fare Reimbursement')

    def _get_service_charge_product(self):
        return self._get_or_create_product('product_service_charge', 'Travel Service Charge')

    def _get_or_create_gst_taxes(self):
        """Get or create CGST 9% and SGST 9% sale taxes."""
        company = self.env.company
        cgst = self.env['account.tax'].search([
            ('name', '=', 'CGST 9%'),
            ('company_id', '=', company.id),
            ('type_tax_use', '=', 'sale'),
        ], limit=1)
        if not cgst:
            cgst = self.env['account.tax'].create({
                'name': 'CGST 9%',
                'type_tax_use': 'sale',
                'amount_type': 'percent',
                'amount': 9.0,
                'company_id': company.id,
                'description': 'CGST @ 9%',
            })
        sgst = self.env['account.tax'].search([
            ('name', '=', 'SGST 9%'),
            ('company_id', '=', company.id),
            ('type_tax_use', '=', 'sale'),
        ], limit=1)
        if not sgst:
            sgst = self.env['account.tax'].create({
                'name': 'SGST 9%',
                'type_tax_use': 'sale',
                'amount_type': 'percent',
                'amount': 9.0,
                'company_id': company.id,
                'description': 'SGST @ 9%',
            })
        return cgst, sgst

    def _get_or_create_igst_tax(self):
        """Get or create IGST 18% sale tax for inter-state transactions."""
        company = self.env.company
        igst = self.env['account.tax'].search([
            ('name', '=', 'IGST 18%'),
            ('company_id', '=', company.id),
            ('type_tax_use', '=', 'sale'),
        ], limit=1)
        if not igst:
            igst = self.env['account.tax'].create({
                'name': 'IGST 18%',
                'type_tax_use': 'sale',
                'amount_type': 'percent',
                'amount': 18.0,
                'company_id': company.id,
                'description': 'IGST @ 18%',
            })
        return igst

    def _get_service_charge_taxes(self, customer):
        """Return the correct GST taxes based on customer state vs company state.
        Intra-state: CGST 9% + SGST 9%
        Inter-state: IGST 18%
        """
        company_state = self.env.company.state_id
        customer_state = customer.state_id if customer else False
        if company_state and customer_state and company_state != customer_state:
            # Inter-state → IGST
            igst = self._get_or_create_igst_tax()
            return igst
        else:
            # Intra-state (or state unknown) → CGST + SGST
            cgst, sgst = self._get_or_create_gst_taxes()
            return cgst | sgst

    def _get_or_create_purchase_gst_taxes(self):
        """Get or create CGST 9% and SGST 9% taxes for purchase."""
        company = self.env.company
        cgst = self.env['account.tax'].search([
            ('name', '=', 'CGST 9%'),
            ('company_id', '=', company.id),
            ('type_tax_use', '=', 'purchase'),
        ], limit=1)
        if not cgst:
            cgst = self.env['account.tax'].create({
                'name': 'CGST 9%',
                'type_tax_use': 'purchase',
                'amount_type': 'percent',
                'amount': 9.0,
                'company_id': company.id,
                'description': 'CGST @ 9%',
            })
        sgst = self.env['account.tax'].search([
            ('name', '=', 'SGST 9%'),
            ('company_id', '=', company.id),
            ('type_tax_use', '=', 'purchase'),
        ], limit=1)
        if not sgst:
            sgst = self.env['account.tax'].create({
                'name': 'SGST 9%',
                'type_tax_use': 'purchase',
                'amount_type': 'percent',
                'amount': 9.0,
                'company_id': company.id,
                'description': 'SGST @ 9%',
            })
        return cgst, sgst

    def _create_draft_sale_order(self):
        """Create a Draft Sales Order with ticket cost + service charge lines."""
        self.ensure_one()
        customer = self._get_customer_partner()
        if not customer:
            return False

        ticket_amount = self._get_ticket_amount()
        sc_amount = self._get_service_charge_amount()
        if not ticket_amount and not sc_amount:
            return False

        booking_type = self._get_booking_type_label()
        passenger = self._get_passenger_info()
        pnr = self._get_pnr_info()

        # Build description lines
        ticket_desc = f"{booking_type} Fare Reimbursement As Per Attached Annex."

        sc_desc = f"Service Charges for {booking_type} Booking"

        company_currency = self.env.company.currency_id.id
        fare_product = self._get_fare_product()
        sc_product = self._get_service_charge_product()

        order_lines = []
        # Line 1: Ticket cost — non-taxable
        if ticket_amount:
            order_lines.append((0, 0, {
                'product_id': fare_product.id,
                'name': ticket_desc,
                'product_uom_qty': 1,
                'price_unit': ticket_amount,
                'tax_id': [(5, 0, 0)],  # No tax
                'currency_id': company_currency,
            }))

        # Line 2: Service charge — taxable (CGST+SGST or IGST based on state)
        if sc_amount:
            taxes = self._get_service_charge_taxes(customer)
            order_lines.append((0, 0, {
                'product_id': sc_product.id,
                'name': sc_desc,
                'product_uom_qty': 1,
                'price_unit': sc_amount,
                'tax_id': [(6, 0, taxes.ids)],
                'currency_id': company_currency,
            }))

        so = self.env['sale.order'].create({
            'partner_id': customer.id,
            'currency_id': company_currency,
            'origin': self.name,
            'name': self.name,
            'note': f"Auto-generated from {booking_type} booking {self.name}",
            'order_line': order_lines,
        })
        self.sale_order_id = so.id
        return so

    def _create_draft_purchase_order(self):
        """Create a Draft Purchase Order mapped to the vendor."""
        self.ensure_one()
        vendor = self._get_vendor_partner()
        if not vendor:
            return False

        ticket_amount = self._get_ticket_amount()
        if not ticket_amount:
            return False

        booking_type = self._get_booking_type_label()
        passenger = self._get_passenger_info()
        pnr = self._get_pnr_info()

        line_desc = f"{booking_type} - {self.name}"
        if pnr:
            line_desc += f" | PNR: {pnr}"
        if passenger:
            line_desc += f" | {passenger}"

        company_currency = self.env.company.currency_id.id
        fare_product = self._get_fare_product()
        po = self.env['purchase.order'].create({
            'partner_id': vendor.id,
            'currency_id': company_currency,
            'origin': self.name,
            'notes': f"Auto-generated from {booking_type} booking {self.name}",
            'order_line': [(0, 0, {
                'product_id': fare_product.id,
                'name': line_desc,
                'product_qty': 1,
                'price_unit': ticket_amount,
                'product_uom': self.env.ref('uom.product_uom_unit').id,
                'currency_id': company_currency,
                'date_planned': fields.Datetime.now(),
                'taxes_id': [(5, 0, 0)],  # No tax on purchase cost
            })],
        })
        self.purchase_order_id = po.id
        return po

    def _generate_sale_purchase_orders(self):
        """Called from action_confirm to generate both SO and PO.
        Auto-confirms SO, then creates TWO separate invoices:
          1) Tax Invoice  — service charge line only (with CGST+SGST)
          2) Debit Note   — fare line only (no tax)
        """
        for rec in self:
            if not rec.sale_order_id:
                so = rec._create_draft_sale_order()
                if so:
                    rec._attach_annexure_to_so(so)
                    # SO stays in draft — senior staff will confirm/post
                    rec._create_split_invoices(so)
            if not rec.purchase_order_id:
                rec._create_draft_purchase_order()

    def _create_split_invoices(self, so):
        """Create two invoices from one SO:
        Invoice 1 (Tax Invoice): service charge line only
        Invoice 2 (Debit Note): fare line only
        """
        sc_lines = so.order_line.filtered(
            lambda l: l.product_id.default_code == 'product_service_charge'
        )
        fare_lines = so.order_line.filtered(
            lambda l: l.product_id.default_code == 'product_travel_fare'
        )

        # Invoice 1: Service Charge (Tax Invoice) — stays in draft
        if sc_lines:
            inv1 = self._create_invoice_from_so_lines(so, sc_lines)

        # Invoice 2: Fare (Debit Note) — stays in draft
        if fare_lines:
            inv2 = self._create_invoice_from_so_lines(so, fare_lines)

    def _create_invoice_from_so_lines(self, so, so_lines):
        """Create a single draft invoice from specific SO lines.
        Builds invoice lines directly from SO line data so the SO
        can remain in draft state.
        """
        invoice_line_vals = []
        for line in so_lines:
            invoice_line_vals.append((0, 0, {
                'product_id': line.product_id.id,
                'name': line.name,
                'quantity': line.product_uom_qty,
                'price_unit': line.price_unit,
                'tax_ids': [(6, 0, line.tax_id.ids)],
            }))

        if not invoice_line_vals:
            return False

        invoice = self.env['account.move'].sudo().create({
            'move_type': 'out_invoice',
            'partner_id': so.partner_id.id,
            'invoice_origin': so.name,
            'invoice_line_ids': invoice_line_vals,
        })

        # Link invoice lines back to SO lines
        for inv_line in invoice.invoice_line_ids.filtered(lambda l: l.display_type == 'product'):
            matching_so_line = so_lines.filtered(
                lambda sl: sl.product_id == inv_line.product_id
            )[:1]
            if matching_so_line:
                inv_line.write({'sale_line_ids': [(4, matching_so_line.id)]})

        return invoice

    def _get_annexure_type_columns(self, booking_type, annexure_kind):
        """Return (headers, values) specific to each booking type for annexures.
        annexure_kind: 'sc' for service charge, 'fare' for fare charge.
        """
        # Common data
        pax = self._get_passenger_info()
        pax_list = self._get_passenger_list()
        num_pax = len(pax_list) if pax_list else 1
        emp_code = getattr(self, 'employee_code', '') or ''
        doc_no = getattr(self, 'document_number', '') or ''
        booking_date_str = str(self.booking_date) if hasattr(self, 'booking_date') and self.booking_date else ''
        pnr = self._get_pnr_info()

        # Common prefix columns for all types
        common_headers = ['SR No.', 'Booking Date', 'No of Pax.', 'Passenger(s) Name', 'Emp. Code', 'Doc. No.']
        common_vals = [1, booking_date_str, num_pax, pax, emp_code, doc_no]

        # Type-specific middle columns
        mid_headers = []
        mid_vals = []

        if booking_type in ('Domestic Flight', 'International Flight'):
            route = self._get_route_info()
            origin = route.split(' → ')[0] if ' → ' in route else route
            dest = route.split(' → ')[1] if ' → ' in route else ''
            trip_type = ''
            if hasattr(self, 'trip_type') and self.trip_type:
                trip_type = dict(self._fields['trip_type'].selection).get(self.trip_type, '')
            dates = self._get_travel_dates_info()
            travel_date = str(dates[0]['date']) if dates else ''
            return_date = str(dates[1]['date']) if len(dates) > 1 else ''
            ticket = getattr(self, 'ticket_number', '') or ''
            flight_no = getattr(self, 'flight_number', '') or ''
            flight_ret = getattr(self, 'flight_number_return', '') or ''
            travel_class = ''
            if hasattr(self, 'travel_class') and self.travel_class:
                travel_class = dict(self._fields['travel_class'].selection).get(self.travel_class, '')
            airline = ''
            if hasattr(self, 'airline') and self.airline:
                field = self._fields.get('airline')
                if field and field.type == 'many2one':
                    airline = self.airline.name or ''
                elif field and field.type == 'selection':
                    airline = dict(field.selection).get(self.airline, self.airline)
                else:
                    airline = self.airline
            if annexure_kind == 'sc':
                mid_headers = ['From Origin', 'To Destination', 'Travel Type', 'Travel Date',
                               'Return Date', 'PNR / Ticket', 'PNR No.', 'Flight No.', 'Class', 'Airline']
                mid_vals = [origin, dest, trip_type, travel_date, return_date,
                            ticket or pnr, pnr, flight_no, travel_class, airline]
            else:
                mid_headers = ['From Origin', 'To Destination', 'Travel Type', 'Travel Date',
                               'Return Date', 'PNR / Ticket', 'PNR No.',
                               'Flight No. Onward', 'Flight No. Return', 'Class', 'Airline']
                mid_vals = [origin, dest, trip_type, travel_date, return_date,
                            ticket or pnr, pnr, flight_no, flight_ret, travel_class, airline]

        elif booking_type == 'Train':
            route = self._get_route_info()
            origin = route.split(' → ')[0] if ' → ' in route else route
            dest = route.split(' → ')[1] if ' → ' in route else ''
            dates = self._get_travel_dates_info()
            travel_date = str(dates[0]['date']) if dates else ''
            train_no = getattr(self, 'train_number', '') or ''
            travel_class = ''
            if hasattr(self, 'travel_class') and self.travel_class:
                travel_class = dict(self._fields['travel_class'].selection).get(self.travel_class, '')
            quota = ''
            if hasattr(self, 'quota') and self.quota:
                quota = dict(self._fields['quota'].selection).get(self.quota, '')
            mid_headers = ['From Origin', 'To Destination', 'Travel Date', 'PNR No.',
                           'Train No.', 'Class', 'Quota']
            mid_vals = [origin, dest, travel_date, pnr, train_no, travel_class, quota]

        elif booking_type == 'Bus':
            route = self._get_route_info()
            origin = route.split(' → ')[0] if ' → ' in route else route
            dest = route.split(' → ')[1] if ' → ' in route else ''
            dates = self._get_travel_dates_info()
            travel_date = str(dates[0]['date']) if dates else ''
            bus_operator = getattr(self, 'bus_operator', '') or ''
            mid_headers = ['From Origin', 'To Destination', 'Travel Date', 'PNR / Ref No.', 'Bus Operator']
            mid_vals = [origin, dest, travel_date, pnr, bus_operator]

        elif booking_type == 'Hotel':
            hotel_name = getattr(self, 'hotel_name', '') or ''
            loc_field = getattr(self, 'location', False)
            location = loc_field.name if loc_field and hasattr(loc_field, 'name') else ''
            if not isinstance(location, str):
                location = str(location) if location else ''
            checkin = str(self.checkin_date) if hasattr(self, 'checkin_date') and self.checkin_date else ''
            checkout = str(self.checkout_date) if hasattr(self, 'checkout_date') and self.checkout_date else ''
            num_rooms = getattr(self, 'num_rooms', '') or ''
            room_type = ''
            if hasattr(self, 'room_type') and self.room_type:
                field = self._fields.get('room_type')
                if field and field.type == 'selection':
                    room_type = dict(field.selection).get(self.room_type, self.room_type)
                else:
                    room_type = self.room_type
            mid_headers = ['Hotel Name', 'Location', 'Check-in', 'Check-out', 'No. of Rooms', 'Room Type']
            mid_vals = [hotel_name, location, checkin, checkout, num_rooms, room_type]

        elif booking_type == 'Car':
            pickup_loc = getattr(self, 'pickup_location', '') or ''
            drop_loc = getattr(self, 'drop_location', '') or ''
            pickup_dt = str(self.pickup_date) if hasattr(self, 'pickup_date') and self.pickup_date else ''
            drop_dt = str(self.drop_date) if hasattr(self, 'drop_date') and self.drop_date else ''
            car_type = ''
            if hasattr(self, 'car_type') and self.car_type:
                car_type = dict(self._fields['car_type'].selection).get(self.car_type, '')
            rental_type = ''
            if hasattr(self, 'rental_type') and self.rental_type:
                rental_type = dict(self._fields['rental_type'].selection).get(self.rental_type, '')
            cab_vendor = getattr(self, 'cab_vendor', '') or ''
            mid_headers = ['Pickup Location', 'Drop Location', 'Pickup Date', 'Drop Date',
                           'Car Type', 'Rental Type', 'Cab Vendor']
            mid_vals = [pickup_loc, drop_loc, pickup_dt, drop_dt, car_type, rental_type, cab_vendor]

        elif booking_type == 'Event':
            event_name = getattr(self, 'event_name', '') or ''
            event_date = str(self.event_date) if hasattr(self, 'event_date') and self.event_date else ''
            event_loc = getattr(self, 'event_location', '') or ''
            mid_headers = ['Event Name', 'Event Date', 'Event Location']
            mid_vals = [event_name, event_date, event_loc]

        elif booking_type == 'Insurance':
            ref_no = getattr(self, 'reference_number', '') or ''
            desc = getattr(self, 'description', '') or ''
            mid_headers = ['Reference No.', 'Description']
            mid_vals = [ref_no, desc]

        elif booking_type == 'Visa':
            btype = ''
            if hasattr(self, 'booking_type') and self.booking_type:
                btype = dict(self._fields['booking_type'].selection).get(self.booking_type, '')
            ref_no = getattr(self, 'reference_number', '') or ''
            desc = getattr(self, 'description', '') or ''
            mid_headers = ['Type', 'Reference No.', 'Description']
            mid_vals = [btype, ref_no, desc]

        elif booking_type == 'Package Tour':
            btype = ''
            if hasattr(self, 'booking_type') and self.booking_type:
                btype = dict(self._fields['booking_type'].selection).get(self.booking_type, '')
            loc_type = ''
            if hasattr(self, 'location_type') and self.location_type:
                loc_type = dict(self._fields['location_type'].selection).get(self.location_type, '')
            ref_no = getattr(self, 'reference_number', '') or ''
            desc = getattr(self, 'description', '') or ''
            mid_headers = ['Type', 'Location', 'Reference No.', 'Description']
            mid_vals = [btype, loc_type, ref_no, desc]

        else:
            # Fallback
            dates = self._get_travel_dates_info()
            travel_date = str(dates[0]['date']) if dates else ''
            mid_headers = ['Travel Date', 'Ref No.']
            mid_vals = [travel_date, pnr]

        return common_headers, common_vals, mid_headers, mid_vals

    def _attach_annexure_to_so(self, so):
        """Generate detailed Excel annexures (service charge + fare) and attach to SO."""
        self.ensure_one()
        try:
            import base64
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        header_font = Font(bold=True, color='FFFFFF', size=9)
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True, size=10)

        booking_type = self._get_booking_type_label()
        customer = self._get_customer_partner()
        customer_name = customer.name if customer else ''
        total_fare = self._get_ticket_amount()
        sc = self._get_service_charge_amount()

        # --- Service Charge Annexure ---
        if sc > 0:
            common_h, common_v, mid_h, mid_v = self._get_annexure_type_columns(booking_type, 'sc')
            headers = common_h + mid_h + ['Service Charge', 'CGST 9%', 'SGST 9%', 'IGST', 'Total ₹', 'Remark']

            cgst_amt = round(sc * 0.09, 2)
            sgst_amt = round(sc * 0.09, 2)
            line_total = round(sc + cgst_amt + sgst_amt, 2)
            vals = common_v + mid_v + [sc, cgst_amt, sgst_amt, 0, line_total, '']

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Service Charge'
            ws.cell(row=1, column=1, value='THE EARTH').font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f'GSTIN: {self.env.company.vat or ""}').font = Font(size=9)
            ws.cell(row=3, column=1, value=f'ANNEXURE - SERVICE CHARGE - {booking_type.upper()} BOOKING').font = Font(bold=True, size=11)
            ws.cell(row=4, column=1, value=f'To: {customer_name}').font = bold_font

            row = 6
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            row += 1
            for col, v in enumerate(vals, 1):
                # Convert Odoo recordsets to display name for Excel compatibility
                if hasattr(v, '_name'):
                    v = v.display_name or '' if v else ''
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border

            row += 1
            sc_col = headers.index('Service Charge') + 1
            ws.cell(row=row, column=3, value='Grand Total').font = bold_font
            for offset, val in enumerate([sc, cgst_amt, sgst_amt, 0, line_total]):
                cell = ws.cell(row=row, column=sc_col + offset, value=val)
                cell.font = bold_font
                cell.border = thin_border

            for c in ws.columns:
                ml = max(len(str(cell.value or '')) for cell in c)
                ws.column_dimensions[c[0].column_letter].width = min(ml + 2, 30)

            fp = io.BytesIO()
            wb.save(fp)
            fp.seek(0)
            self.env['ir.attachment'].create({
                'name': f"Annexure_ServiceCharge_{self.name}.xlsx",
                'type': 'binary',
                'datas': base64.b64encode(fp.read()),
                'res_model': 'sale.order',
                'res_id': so.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            fp.close()

        # --- Fare Charge Annexure ---
        if total_fare > 0:
            common_h, common_v, mid_h, mid_v = self._get_annexure_type_columns(booking_type, 'fare')
            headers2 = common_h + mid_h + ['Total Fare', 'Cancel Cost', 'Refund Amount', 'Inv No.', 'Remark']
            vals2 = common_v + mid_v + [total_fare, 0, 0, so.name or '', '']

            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = 'Fare Charge'
            ws2.cell(row=1, column=1, value='THE EARTH').font = Font(bold=True, size=14)
            ws2.cell(row=2, column=1, value=f'GSTIN: {self.env.company.vat or ""}').font = Font(size=9)
            ws2.cell(row=3, column=1, value=f'ANNEXURE - FARE CHARGE - {booking_type.upper()} BOOKING').font = Font(bold=True, size=11)
            ws2.cell(row=4, column=1, value=f'To: {customer_name}').font = bold_font

            row = 6
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            row += 1
            for col, v in enumerate(vals2, 1):
                # Convert Odoo recordsets to display name for Excel compatibility
                if hasattr(v, '_name'):
                    v = v.display_name or '' if v else ''
                cell = ws2.cell(row=row, column=col, value=v)
                cell.border = thin_border

            row += 1
            fare_col = headers2.index('Total Fare') + 1
            ws2.cell(row=row, column=3, value='Grand Total').font = bold_font
            ws2.cell(row=row, column=fare_col, value=total_fare).font = bold_font
            ws2.cell(row=row, column=fare_col).border = thin_border

            for c in ws2.columns:
                ml = max(len(str(cell.value or '')) for cell in c)
                ws2.column_dimensions[c[0].column_letter].width = min(ml + 2, 30)

            fp2 = io.BytesIO()
            wb2.save(fp2)
            fp2.seek(0)
            self.env['ir.attachment'].create({
                'name': f"Annexure_FareCharge_{self.name}.xlsx",
                'type': 'binary',
                'datas': base64.b64encode(fp2.read()),
                'res_model': 'sale.order',
                'res_id': so.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            fp2.close()

    def action_view_sale_order(self):
        """Open the linked Sales Order."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Sales Order'),
            'res_model': 'sale.order',
            'res_id': self.sale_order_id.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_view_purchase_order(self):
        """Open the linked Purchase Order."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Purchase Order'),
            'res_model': 'purchase.order',
            'res_id': self.purchase_order_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
