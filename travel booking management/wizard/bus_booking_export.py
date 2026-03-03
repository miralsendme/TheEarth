# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, _

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None


class BusBookingExport(models.TransientModel):
    _name = 'travel.bus.booking.export'
    _description = 'Export Bus Bookings to Excel'

    date_from = fields.Date(string='From Date')
    date_to = fields.Date(string='To Date')
    state_filter = fields.Selection([
        ('all', 'All'),
        ('draft', 'Draft'),
        ('confirmed', 'Confirmed'),
        ('done', 'Done'),
        ('cancelled', 'Cancelled'),
    ], string='Status', default='all')
    excel_file = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='File Name', readonly=True)

    def action_export(self):
        self.ensure_one()
        domain = []
        if self.date_from:
            domain.append(('booking_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('booking_date', '<=', self.date_to))
        if self.state_filter and self.state_filter != 'all':
            domain.append(('state', '=', self.state_filter))

        if self.env.context.get('active_ids') and self.env.context.get('active_model') == 'travel.bus.booking':
            domain = [('id', 'in', self.env.context['active_ids'])]
            if self.state_filter and self.state_filter != 'all':
                domain.append(('state', '=', self.state_filter))

        bookings = self.env['travel.bus.booking'].search(domain, order='booking_date asc, departure_date asc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bus Bookings'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = [
            'Booking Ref', 'Billing Company', 'Booking Date', 'Booking Executive',
            'Employee Code', 'Document No / Requested By', 'Bus Operator', 'Bus Type',
            'Boarding Point', 'Dropping Point', 'Departure', 'Arrival',
            'Passenger(s)', 'No. of Passengers', 'Seat Numbers', 'PNR Number',
            'Amount', 'Mode of Payment', 'Confirmed By', 'Status', 'Notes',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, bk in enumerate(bookings, 2):
            values = [
                bk.name or '',
                bk.billing_company_id.name or '',
                str(bk.booking_date) if bk.booking_date else '',
                bk.booking_executive.name if bk.booking_executive else '',
                bk.employee_code or '',
                bk.document_number or '',
                bk.operator_name or '',
                dict(bk._fields['bus_type'].selection).get(bk.bus_type, ''),
                bk.boarding_point or '',
                bk.dropping_point or '',
                str(bk.departure_date) if bk.departure_date else '',
                str(bk.arrival_date) if bk.arrival_date else '',
                (bk.passenger_names or '').replace('\n', ', '),
                bk.num_passengers or 0,
                bk.seat_numbers or '',
                bk.pnr_number or '',
                bk.total_amount or 0.0,
                dict(bk._fields['mode_of_payment'].selection).get(bk.mode_of_payment, '') if bk.mode_of_payment else '',
                bk.confirmed_by or '',
                dict(bk._fields['state'].selection).get(bk.state, ''),
                bk.notes or '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        self.write({
            'excel_file': base64.b64encode(fp.read()),
            'file_name': 'bus_bookings.xlsx',
        })
        fp.close()

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
